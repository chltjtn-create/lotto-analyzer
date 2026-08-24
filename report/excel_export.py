"""Excel report export for Lotto analysis results."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from lotto_analyzer.analysis.backtest import BacktestSummary
from lotto_analyzer.analysis.evaluation import RecommendationEvaluation, RecommendationRecord
from lotto_analyzer.analysis.frequency import NumberFrequencyStats, stats_to_rows
from lotto_analyzer.analysis.pattern import PatternSummary, pattern_rows
from lotto_analyzer.analysis.scoring import NumberScore, score_rows
from lotto_analyzer.config import BASE_DIR
from lotto_analyzer.domain.models import LottoDraw


class ExcelExportError(Exception):
    """Raised when the Excel report cannot be exported."""


def export_excel_report(
    draws: Iterable[LottoDraw],
    stats_by_number: dict[int, NumberFrequencyStats],
    pattern_summary: PatternSummary,
    scores_by_number: dict[int, NumberScore],
    recommendations: Iterable[RecommendationRecord] | None = None,
    evaluations: Iterable[RecommendationEvaluation] | None = None,
    backtest_summary: BacktestSummary | None = None,
    output_path: Path | str = BASE_DIR / "lotto_report.xlsx",
) -> Path:
    """Export the main MVP report workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ExcelExportError("openpyxl is required to export Excel reports.") from exc

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    draw_list = sorted(draws, key=lambda draw: draw.draw_no)
    _write_rows(
        wb.create_sheet("데이터요약"),
        [draw.to_dict() for draw in draw_list],
        Font,
    )
    _write_rows(wb.create_sheet("통계분석"), stats_to_rows(stats_by_number), Font)
    _write_rows(wb.create_sheet("패턴분석"), pattern_rows(pattern_summary), Font)
    _write_rows(wb.create_sheet("번호점수"), score_rows(scores_by_number), Font)
    _write_rows(
        wb.create_sheet("추천조합"),
        [record.to_dict() for record in recommendations or []],
        Font,
    )
    _write_rows(
        wb.create_sheet("추천평가"),
        [evaluation.to_dict() for evaluation in evaluations or []],
        Font,
    )
    backtest_rows = [round_result.to_dict() for round_result in backtest_summary.rounds] if backtest_summary else []
    if backtest_summary:
        backtest_rows.insert(0, backtest_summary.to_dict())
    _write_rows(wb.create_sheet("백테스트"), backtest_rows, Font)

    try:
        wb.save(path)
    except OSError as exc:
        raise ExcelExportError(f"Failed to save Excel report: {exc}") from exc
    return path


def _write_rows(sheet, rows: list[dict[str, object]], font_cls) -> None:
    """Write dictionaries to a worksheet with a bold header row."""
    if not rows:
        sheet.append(["데이터 없음"])
        sheet["A1"].font = font_cls(bold=True)
        return

    headers = list(rows[0])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = font_cls(bold=True)

    for row in rows:
        sheet.append([_stringify_cell(row.get(header)) for header in headers])

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 45)


def _stringify_cell(value: object) -> object:
    """Convert list-like cell values into readable text."""
    if isinstance(value, list | tuple):
        return ", ".join(str(item) for item in value)
    return value
