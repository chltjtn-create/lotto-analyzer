"""Load Lotto draw data from local CSV/XLSX files."""

from __future__ import annotations

from datetime import date, datetime
from datetime import timedelta
from pathlib import Path
from typing import Iterable

from lotto_analyzer.config import FIRST_LOTTO_DRAW_DATE
from lotto_analyzer.domain.models import LottoDraw
from lotto_analyzer.domain.validators import DrawValidationError


class LocalDataLoadError(Exception):
    """Raised when local Lotto data cannot be loaded."""


def load_draws_from_excel(path: Path | str, sheet_name: str | None = None) -> list[LottoDraw]:
    """Load draw data from the local Excel workbook in the workspace."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise LocalDataLoadError("openpyxl is required to import Excel data.") from exc

    excel_path = Path(path)
    if not excel_path.exists():
        raise LocalDataLoadError(f"Excel file not found: {excel_path}")

    workbook = None
    try:
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    except Exception as exc:
        raise LocalDataLoadError(f"Failed to open Excel file: {exc}") from exc

    draws: list[LottoDraw] = []
    try:
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            parsed = _parse_excel_row(row, row_index)
            if parsed is not None:
                draws.append(parsed)
    finally:
        if workbook is not None:
            workbook.close()

    if not draws:
        raise LocalDataLoadError("No valid draw rows were found in the Excel file.")

    return sorted(draws, key=lambda draw: draw.draw_no)


def _parse_excel_row(row: Iterable[object], row_index: int) -> LottoDraw | None:
    """Parse one workbook row, skipping headers and non-data rows."""
    values = list(row)
    if _looks_like_official_download_row(values):
        return _parse_official_download_row(values, row_index)
    if _looks_like_legacy_workspace_row(values):
        return _parse_legacy_workspace_row(values, row_index)
    return None


def _looks_like_official_download_row(values: list[object]) -> bool:
    """Return whether a row matches the current official download layout."""
    if len(values) < 9:
        return False
    draw_no = _try_parse_int(values[1])
    return draw_no is not None and draw_no >= 1


def _looks_like_legacy_workspace_row(values: list[object]) -> bool:
    """Return whether a row matches the previous workspace workbook layout."""
    if len(values) < 19:
        return False
    draw_no = _try_parse_int(values[0])
    return draw_no is not None and draw_no >= 1


def _parse_official_download_row(values: list[object], row_index: int) -> LottoDraw:
    """Parse the official result download layout without a draw-date column."""
    draw_no = _parse_int(values[1], row_index, "draw_no")
    draw_date = FIRST_LOTTO_DRAW_DATE + timedelta(days=(draw_no - 1) * 7)
    numbers = tuple(_parse_int(values[index], row_index, f"num{index - 1}") for index in range(2, 8))
    bonus = _parse_int(values[8], row_index, "bonus")
    return _build_draw(draw_no, draw_date, numbers, bonus, row_index)


def _parse_legacy_workspace_row(values: list[object], row_index: int) -> LottoDraw:
    """Parse the previous workspace workbook layout with an explicit draw date."""
    draw_no = _parse_int(values[0], row_index, "draw_no")
    draw_date = _parse_date(values[1], row_index)
    numbers = tuple(_parse_int(values[index], row_index, f"num{index - 11}") for index in range(12, 18))
    bonus = _parse_int(values[18], row_index, "bonus")
    return _build_draw(draw_no, draw_date, numbers, bonus, row_index)


def _build_draw(
    draw_no: int,
    draw_date: date,
    numbers: tuple[int, ...],
    bonus: int,
    row_index: int,
) -> LottoDraw:
    """Build a validated LottoDraw and wrap validation errors with row context."""
    try:
        return LottoDraw(draw_no=draw_no, draw_date=draw_date, numbers=numbers, bonus=bonus)
    except DrawValidationError as exc:
        raise LocalDataLoadError(f"Invalid draw data at row {row_index}: {exc}") from exc


def _try_parse_int(value: object) -> int | None:
    """Try parsing an integer value and return None on failure."""
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return None


def _parse_date(value: object, row_index: int) -> date:
    """Parse a date cell from Excel into a date object."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y.%m.%d", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    raise LocalDataLoadError(f"Invalid draw date at row {row_index}: {value}")


def _parse_int(value: object, row_index: int, field_name: str) -> int:
    """Parse an integer cell from Excel."""
    parsed = _try_parse_int(value)
    if parsed is None:
        raise LocalDataLoadError(f"Invalid {field_name} at row {row_index}: {value}")
    return parsed
